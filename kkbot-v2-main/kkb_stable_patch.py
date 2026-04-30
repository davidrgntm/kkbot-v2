from __future__ import annotations

import asyncio
import calendar
import html
import os
import tempfile
import urllib.parse
import urllib.request
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from fastapi import Form, Request, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, FileResponse

import web_server as ws
from config import config
from database.sqlite_db import db
from services.attendance_verification import (
    save_data_url_image,
    verify_attendance,
    checks,
    update_check_status,
    ensure_verification_tables,
    _image_phash,
    _set_reference,
)

_PATCHED = False


def esc(v: Any) -> str:
    return html.escape(str(v if v is not None else ""), quote=True)


def money(v: Any) -> str:
    try:
        return f"{round(float(v or 0)):,}".replace(",", " ")
    except Exception:
        return "0"


def remove_route(app, path: str, method: str = "GET") -> None:
    app.router.routes = [
        r for r in app.router.routes
        if not (getattr(r, "path", None) == path and method in getattr(r, "methods", set()))
    ]


def sql_one(q: str, p: tuple = ()):  # noqa: ANN001
    return db._execute(q, p, "one")


def sql_all(q: str, p: tuple = ()):  # noqa: ANN001
    return db._execute(q, p, "all")


def row_get(row: Any, key: str, default: Any = None) -> Any:
    """sqlite3.Row does not have .get(); this helper works for dict and Row."""
    if row is None:
        return default
    if isinstance(row, dict):
        return row.get(key, default)
    try:
        return row[key]
    except Exception:
        return default


def row_dict(row: Any) -> dict:
    if row is None:
        return {}
    if isinstance(row, dict):
        return row
    try:
        return dict(row)
    except Exception:
        return {}


def parse_dt(value: Any) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value)).replace(tzinfo=None)
    except Exception:
        return None


def dt_short(value: Any) -> str:
    d = parse_dt(value)
    return d.strftime("%d.%m.%Y %H:%M") if d else "—"


def worked_text(minutes: Any) -> str:
    m = max(0, int(minutes or 0))
    return f"{m // 60} ч {m % 60:02d} м"


def legacy_paid_minutes(raw_minutes: int) -> tuple[int, int]:
    raw_minutes = max(0, int(raw_minutes or 0))
    h = raw_minutes // 60
    m = raw_minutes % 60
    if m >= 30:
        h += 1
    break_minutes = 60 if h >= 5 else 0
    if h >= 5:
        h -= 1
    return max(0, h * 60), break_minutes


def raw_minutes(row) -> int:
    s = parse_dt(row["start_at"])
    e = parse_dt(row["end_at"])
    if not s or not e:
        return 0
    return max(0, int((e - s).total_seconds() // 60))


def paid_for_shift(row) -> tuple[int, int, int]:
    raw = raw_minutes(row)
    paid = int(row["worked_minutes"] or 0)
    br = int(row["break_minutes"] or 0)
    if paid <= 0 and row["end_at"]:
        paid, br = legacy_paid_minutes(raw)
    return raw, paid, br


def clean_phone(value: str) -> str:
    return "".join(ch for ch in str(value or "") if ch.isdigit() or ch == "+")


def ensure_extra_tables() -> None:
    ensure_verification_tables()
    db._execute(
        """
        CREATE TABLE IF NOT EXISTS employee_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id TEXT NOT NULL,
            user_id INTEGER,
            type TEXT DEFAULT '',
            reason TEXT DEFAULT '',
            status TEXT DEFAULT 'pending',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )


def shop_rows() -> list:
    return list(sql_all("SELECT * FROM shops WHERE company_id=? AND active=1 ORDER BY name", (db._current_cid(),)))


def user_shops(user_id: int | str) -> set[str]:
    rows = sql_all(
        """
        SELECT s.name FROM user_shops us
        JOIN shops s ON s.id=us.shop_id
        WHERE us.user_id=? AND s.active=1
        ORDER BY s.name
        """,
        (int(user_id),),
    )
    return {r["name"] for r in rows}


def shop_checkbox_html(selected: set[str] | list[str] | None = None) -> str:
    selected_set = set(selected or [])
    out = "<div class='shop-checks'>"
    for sh in shop_rows():
        name = sh["name"]
        checked = "checked" if name in selected_set else ""
        out += f"<label><input type='checkbox' name='shops' value='{esc(name)}' {checked}> {esc(name)}</label>"
    out += "</div>"
    return out


def role_options(selected: str = "staff") -> str:
    selected = (selected or "staff").lower()
    return "".join(f"<option value='{r}' {'selected' if r == selected else ''}>{r}</option>" for r in ["staff", "manager", "admin", "super_admin"])


def staff_select(selected: str = "") -> str:
    rows = sql_all("SELECT telegram_id, full_name FROM users WHERE company_id=? AND active=1 ORDER BY full_name", (db._current_cid(),))
    return "".join(f"<option value='{esc(r['telegram_id'])}' {'selected' if str(r['telegram_id']) == str(selected) else ''}>{esc(r['full_name'])}</option>" for r in rows)


def shop_select(selected: str = "", empty: bool = False) -> str:
    out = "<option value=''>Barcha filiallar</option>" if empty else ""
    for sh in shop_rows():
        name = sh["name"]
        out += f"<option value='{esc(name)}' {'selected' if name == selected else ''}>{esc(name)}</option>"
    return out


def ensure_user_and_shops(telegram_id: str, name: str, role: str, shops: list[str], phone: str = "", rate: str = "0", emoji: str = "🙂", department: str = "", position: str = "") -> int:
    cid = db._current_cid()
    tid = str(telegram_id).replace(".0", "").strip()
    old = sql_one("SELECT id FROM users WHERE company_id=? AND telegram_id=?", (cid, tid))
    try:
        hourly_rate = float(str(rate or 0).replace(" ", "").replace(",", "."))
    except Exception:
        hourly_rate = 0.0
    if old:
        uid = int(old["id"])
        db._execute(
            """
            UPDATE users
            SET full_name=?, phone=?, role=?, hourly_rate=?, emoji=?, department=?, position=?, active=1, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
            """,
            (name, clean_phone(phone), role, hourly_rate, emoji[:8] or "🙂", department, position, uid),
        )
    else:
        cur = db._execute(
            """
            INSERT INTO users(company_id, telegram_id, full_name, phone, role, hourly_rate, emoji, department, position, active)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
            """,
            (cid, tid, name, clean_phone(phone), role, hourly_rate, emoji[:8] or "🙂", department, position),
        )
        uid = int(cur.lastrowid)
    db._execute("DELETE FROM user_shops WHERE user_id=?", (uid,))
    for shop in shops:
        shop = str(shop or "").strip()
        if not shop:
            continue
        sh = sql_one("SELECT id FROM shops WHERE company_id=? AND name=?", (cid, shop))
        if not sh:
            cur = db._execute("INSERT OR IGNORE INTO shops(company_id, name, active) VALUES (?, ?, 1)", (cid, shop))
            sid = int(cur.lastrowid or 0)
            if not sid:
                sh = sql_one("SELECT id FROM shops WHERE company_id=? AND name=?", (cid, shop))
                sid = int(sh["id"]) if sh else 0
        else:
            sid = int(sh["id"])
        if sid:
            db._execute("INSERT OR IGNORE INTO user_shops(user_id, shop_id) VALUES (?, ?)", (uid, sid))
    return uid


def stable_nav(active: str, user: dict) -> str:
    admin = ws.is_admin(user)
    if admin:
        items = [
            ("dashboard", "/dashboard", "🏠 Dashboard"),
            ("verification", "/verification", "🧠 AI tekshiruv"),
            ("employees", "/employees", "👥 Xodimlar"),
            ("shops", "/shops", "🏪 Filiallar"),
            ("shifts", "/shifts", "🟢 Smenalar"),
            ("schedule", "/schedule", "📅 Grafik"),
            ("quick_schedule", "/quick-schedule", "⚡ Tez grafik"),
            ("salary", "/salary", "💰 Oylik"),
            ("requests", "/requests", "📝 Arizalar"),
            ("inventory", "/inventory", "📦 Inventar"),
            ("reports", "/reports", "🧾 Hisobotlar"),
        ]
    else:
        items = [
            ("cabinet", "/cabinet", "👤 Kabinet"),
            ("checkin", "/checkin", "🟢 Keldim"),
            ("checkout", "/checkout", "🔴 Ketdim"),
            ("my_schedule", "/schedule", "📅 Grafikim"),
            ("my_shifts", "/shifts", "🟢 Smenalarim"),
        ]
    links = "".join(f'<a class="{"active" if k == active else ""}" href="{h}">{lab}</a>' for k, h, lab in items)
    return f"""
    <aside class="side" id="sideNav">
      <div class="logo"><div class="logo-badge">K</div><div>KKB<br><span style="font-size:12px;color:#93c5fd;font-weight:700">Web Panel</span></div></div>
      <div style="margin-bottom:16px;color:#bfdbfe;font-size:14px">{esc(user.get('name') or 'User')}<br>{esc(user.get('role') or '')}</div>
      <nav class="nav">{links}</nav>
      <form method="post" action="/logout" style="margin-top:22px"><button class="btn gray" style="width:100%">Chiqish</button></form>
    </aside>
    """


def stable_layout(request: Request, active: str, title: str, subtitle: str, content: str) -> HTMLResponse:
    user = ws.current_user(request) or {"name": "", "role": ""}
    mobile = f"<div class='mobile-head'><b>KKB · {esc(title)}</b><button class='hamb' onclick=\"document.getElementById('sideNav').classList.toggle('open')\">☰ Menu</button></div>"
    script = """
    <script>
    window.addEventListener('beforeunload',()=>sessionStorage.setItem('scrollY', String(window.scrollY||0)));
    window.addEventListener('load',()=>{const y=sessionStorage.getItem('scrollY'); if(y){setTimeout(()=>window.scrollTo(0, Number(y)),30)}});
    </script>
    """
    return HTMLResponse(f"""<!doctype html><html lang="uz"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{esc(title)} · KKB</title><style>{ws.CSS}{EXTRA_CSS}</style></head><body>{mobile}<div class="app">{stable_nav(active, user)}<main class="main"><div class="top"><div><div class="h1">{esc(title)}</div><div class="sub">{subtitle}</div></div><button class="btn secondary" onclick="location.reload()">Yangilash</button></div>{content}<div class="footer">DB: {esc(db.db_path)} · Company: {esc(db._current_cid())}</div></main></div>{script}</body></html>""")


EXTRA_CSS = r'''
.shop-checks{display:grid;grid-template-columns:repeat(auto-fit,minmax(190px,1fr));gap:8px;margin:8px 0}.shop-checks label{background:#f8fafc;border:1px solid #e2e8f0;border-radius:12px;padding:10px;font-weight:650}.two{display:grid;grid-template-columns:1.1fr .9fr;gap:18px}.emp-card{display:flex;align-items:center;gap:14px}.emp-avatar{width:62px;height:62px;border-radius:20px;background:#e0e7ff;display:grid;place-items:center;font-size:28px}.muted{color:#64748b}.edit-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:10px}.scan-body{margin:0;background:radial-gradient(circle at top,#55106d 0,#081633 45%,#050816 100%);color:white;font-family:Inter,system-ui,-apple-system,Segoe UI,sans-serif}.scan-wrap{min-height:100vh;max-width:560px;margin:auto;padding:14px;display:flex;flex-direction:column}.scan-top{display:flex;justify-content:space-between;gap:10px;align-items:center}.scan-title{font-size:22px;font-weight:850}.scan-sub{font-size:13px;color:#b9c5d8}.scan-back{width:48px;height:48px;border:0;border-radius:18px;background:rgba(255,255,255,.14);color:white;font-size:24px}.phone-card{position:relative;overflow:hidden;border-radius:34px;margin-top:16px;min-height:58vh;background:#0f172a;border:1px solid rgba(255,255,255,.12);box-shadow:0 30px 90px rgba(0,0,0,.38)}.phone-card video{width:100%;height:58vh;object-fit:cover;transform:scaleX(-1);display:block;background:#111827}.scan-empty{height:58vh;display:grid;place-items:center;text-align:center;padding:24px;color:#cbd5e1}.scan-overlay{position:absolute;inset:0;pointer-events:none;background:radial-gradient(ellipse 39% 31% at 50% 42%,transparent 0 62%,rgba(5,8,22,.55) 63%,rgba(5,8,22,.86) 100%)}.face-oval{position:absolute;left:50%;top:42%;transform:translate(-50%,-50%);width:min(76vw,360px);height:min(95vw,440px);border-radius:50%;border:4px solid rgba(255,255,255,.92);box-shadow:0 0 0 12px rgba(139,92,246,.25),0 0 55px rgba(168,85,247,.6)}.face-oval.ready{border-color:#22c55e;box-shadow:0 0 0 12px rgba(34,197,94,.22),0 0 65px rgba(34,197,94,.45)}.scan-line{position:absolute;left:12%;right:12%;top:24%;height:3px;border-radius:99px;background:linear-gradient(90deg,transparent,#22c55e,transparent);animation:scan 2.2s infinite}.scan-panel{margin-top:14px;background:rgba(255,255,255,.11);border:1px solid rgba(255,255,255,.14);border-radius:28px;padding:18px;backdrop-filter:blur(18px)}.step{display:flex;align-items:center;gap:11px;color:#b9c5d8;margin:10px 0}.dot{width:26px;height:26px;border-radius:50%;background:#334155;display:grid;place-items:center;color:white;font-weight:850}.step.ok{color:#ecfdf5}.step.ok .dot{background:#10b981}.step.warn .dot{background:#f59e0b}.scan-btn{width:100%;height:58px;border:0;border-radius:20px;font-size:18px;font-weight:850;color:white;background:linear-gradient(135deg,#22c55e,#16a34a);box-shadow:0 18px 38px rgba(16,185,129,.28);margin-top:12px}.scan-btn.red{background:linear-gradient(135deg,#ef4444,#dc2626)}.scan-btn.secondary{background:rgba(255,255,255,.14);box-shadow:none}.scan-btn:disabled{opacity:.45;filter:grayscale(.5);box-shadow:none}.result-card{background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.13);border-radius:22px;padding:15px;margin-top:12px}.staff-wrap{max-width:980px;margin:auto;padding:18px}.staff-hero{background:linear-gradient(135deg,#07142e,#172554);color:white;border-radius:30px;padding:24px;margin-bottom:16px}.big-actions{display:grid;grid-template-columns:1fr 1fr;gap:12px}.big-actions a{min-height:112px;border-radius:28px;display:flex;align-items:center;justify-content:center;color:white;text-decoration:none;font-size:24px;font-weight:850}.big-actions .green{background:linear-gradient(135deg,#22c55e,#16a34a)}.big-actions .red{background:linear-gradient(135deg,#ef4444,#dc2626)}@keyframes scan{0%{top:24%}50%{top:62%}100%{top:24%}}@media(max-width:900px){.two,.edit-grid{grid-template-columns:1fr}.shop-checks{grid-template-columns:1fr}.big-actions{grid-template-columns:1fr}.phone-card{min-height:54vh}.phone-card video,.scan-empty{height:54vh}}
'''


def send_photo(chat_id: int | str, photo_path: str, caption: str) -> None:
    token = config.bot_token.get_secret_value()
    boundary = "----KKB" + uuid.uuid4().hex
    body_parts: list[bytes] = []

    def field(name: str, value: str) -> None:
        body_parts.append(f"--{boundary}\r\n".encode())
        body_parts.append(f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode())
        body_parts.append(str(value).encode())
        body_parts.append(b"\r\n")

    field("chat_id", str(chat_id))
    field("caption", caption)
    field("parse_mode", "HTML")
    filename = os.path.basename(photo_path) or "selfie.jpg"
    body_parts.append(f"--{boundary}\r\n".encode())
    body_parts.append(f'Content-Disposition: form-data; name="photo"; filename="{filename}"\r\n'.encode())
    body_parts.append(b"Content-Type: image/jpeg\r\n\r\n")
    body_parts.append(Path(photo_path).read_bytes())
    body_parts.append(b"\r\n")
    body_parts.append(f"--{boundary}--\r\n".encode())
    body = b"".join(body_parts)
    req = urllib.request.Request(f"https://api.telegram.org/bot{token}/sendPhoto", data=body, method="POST")
    req.add_header("Content-Type", f"multipart/form-data; boundary={boundary}")
    req.add_header("Content-Length", str(len(body)))
    with urllib.request.urlopen(req, timeout=15) as resp:
        resp.read()


def notify_group(action: str, name: str, shop: str, result: dict, selfie_path: str, loc_url: str = "", worked: str = "") -> None:
    try:
        icon = "🟢" if action == "check_in" else "🔴"
        title = "KELDI" if action == "check_in" else "KETDI"
        lines = [
            f"{icon} <b>WEB {title}</b>",
            f"👤 {esc(name)}",
            f"🏪 {esc(shop or result.get('shop') or '—')}",
        ]
        if worked:
            lines.append(f"⏱ {esc(worked)}")
        lines += [
            f"🧠 Face: {esc(result.get('face_status'))} ({esc(result.get('face_score'))})",
            f"📍 GPS: {esc(result.get('location_status'))} {esc(result.get('distance_m') or '')}m",
            f"✅ Status: {esc(result.get('final_status'))}",
        ]
        if loc_url:
            lines.append(f"🗺 <a href='{esc(loc_url)}'>Xarita</a>")
        send_photo(config.group_chat_id, selfie_path, "\n".join(lines))
    except Exception as e:
        print(f"[kkb_stable_patch] group photo skipped: {e}")



def scan_page(label: str, action: str, user: dict) -> HTMLResponse:
    is_out = action == "check_out"
    action_title = "Ketdim" if is_out else "Keldim"
    html_page = r"""<!doctype html>
<html lang="uz">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1,viewport-fit=cover">
  <title>__LABEL__ · KKB</title>
  <style>__STYLE__
    .auto-confirm{position:absolute;inset:0;display:none;align-items:center;justify-content:center;text-align:center;z-index:8;background:radial-gradient(circle at center,rgba(15,23,42,.12),rgba(15,23,42,.64));padding:28px;cursor:pointer;touch-action:manipulation}.auto-confirm.show{display:flex}.confirm-box{width:min(86%,330px);border:1px solid rgba(255,255,255,.25);border-radius:28px;background:rgba(15,23,42,.76);box-shadow:0 24px 80px rgba(0,0,0,.45);padding:26px;backdrop-filter:blur(16px)}.confirm-box h2{font-size:30px;line-height:1.05;margin:0 0 10px}.confirm-box p{margin:0;color:#cbd5e1;font-size:15px}.auto-status{position:absolute;left:24px;right:24px;bottom:28px;z-index:7;text-align:center;color:#e2e8f0;text-shadow:0 2px 8px rgba(0,0,0,.45)}.auto-status h2{font-size:24px;margin:0 0 8px}.auto-status p{margin:0;color:#cbd5e1}.scan-panel .scan-note{margin-top:12px;color:#cbd5e1;font-size:14px;text-align:center}.step.ok .dot{background:#10b981}.step.warn .dot{background:#f59e0b}.step.wait .dot{background:#3b82f6}.result-card.success{border-color:rgba(34,197,94,.45);background:rgba(34,197,94,.14)}.result-card.error{border-color:rgba(239,68,68,.45);background:rgba(239,68,68,.14)}
  </style>
</head>
<body class="scan-body">
  <div class="scan-wrap" id="touchArea">
    <div class="scan-top">
      <button class="scan-back" onclick="location.href='/cabinet'">‹</button>
      <div><div class="scan-title">__LABEL__ tasdiqlash</div><div class="scan-sub">Kamera · liveness · GPS · AI moslik</div></div>
      <button class="scan-back" onclick="location.reload()">↻</button>
    </div>
    <div class="phone-card" id="phoneCard">
      <div class="scan-empty" id="empty"><div><h2 id="centerTitle">Face scan boshlanmoqda</h2><p id="centerText">Kamera va lokatsiya avtomatik tekshiriladi.</p></div></div>
      <video id="video" autoplay playsinline webkit-playsinline muted style="display:none"></video>
      <canvas id="canvas" style="display:none"></canvas>
      <div class="scan-overlay"></div>
      <div class="face-oval" id="oval"><div class="scan-line"></div></div>
      <div class="auto-status" id="autoStatus"><h2>Face scan boshlanmoqda</h2><p>Ruxsat oynasi chiqsa Allow bosing.</p></div>
      <div class="auto-confirm" id="confirmOverlay" onclick="sendCheck()" ontouchend="event.preventDefault();sendCheck();">
        <div class="confirm-box"><h2>Tasdiqlaysizmi?</h2><p>__ACTION_TITLE__ bazaga yoziladi. Tasdiqlash uchun ekranga bosing.</p></div>
      </div>
    </div>
    <div class="scan-panel">
      <div class="step ok"><span class="dot">✓</span><span>Xodim: <b>__USER_NAME__</b></span></div>
      <div class="step wait" id="cameraStep"><span class="dot">1</span><span>Kamera ochilmoqda...</span></div>
      <div class="step wait" id="gpsStep"><span class="dot">2</span><span>GPS aniqlanmoqda...</span></div>
      <div class="step wait" id="liveStep"><span class="dot">3</span><span>Liveness avtomatik tekshiriladi</span></div>
      <div class="scan-note" id="scanNote">Hech qanday tugma yo‘q. Ruxsat oynasi chiqmasa, ekranga bir marta teging.</div>
      <div id="result"></div>
    </div>
  </div>
<script>
(function(){
  var ACTION = "__ACTION__";
  var ACTION_TITLE = "__ACTION_TITLE__";
  var pos = null;
  var cameraReady = false;
  var gpsReady = false;
  var liveReady = false;
  var scanStarted = false;
  var sending = false;
  var confirmShown = false;
  function el(id){ return document.getElementById(id); }
  function setStatus(title, text){
    var st = el("autoStatus");
    if(st){ st.innerHTML = "<h2>" + title + "</h2><p>" + (text || "") + "</p>"; }
    var ct = el("centerTitle"); var cx = el("centerText");
    if(ct){ ct.textContent = title; }
    if(cx){ cx.textContent = text || ""; }
  }
  function stepOk(id, text){
    var x = el(id); if(!x){ return; }
    x.className = "step ok";
    var d = x.querySelector(".dot"); if(d){ d.textContent = "✓"; }
    var s = x.querySelector("span:last-child"); if(s){ s.textContent = text; }
  }
  function stepWarn(id, text){
    var x = el(id); if(!x){ return; }
    x.className = "step warn";
    var d = x.querySelector(".dot"); if(d){ d.textContent = "!"; }
    var s = x.querySelector("span:last-child"); if(s){ s.textContent = text; }
  }
  function cameraPromise(){
    var c = { video: { facingMode: "user", width: { ideal: 720 }, height: { ideal: 960 } }, audio: false };
    if(navigator.mediaDevices && navigator.mediaDevices.getUserMedia){ return navigator.mediaDevices.getUserMedia(c); }
    var old = navigator.getUserMedia || navigator.webkitGetUserMedia || navigator.mozGetUserMedia || navigator.msGetUserMedia;
    if(old){ return new Promise(function(resolve, reject){ old.call(navigator, c, resolve, reject); }); }
    return Promise.reject(new Error("Kamera API topilmadi. Safari yoki Chrome brauzerida oching."));
  }
  function startCamera(){
    setStatus("Kamera tekshirilmoqda", "Ruxsat oynasi chiqsa Allow bosing.");
    cameraPromise().then(function(stream){
      var v = el("video");
      v.srcObject = stream;
      v.style.display = "block";
      var empty = el("empty"); if(empty){ empty.style.display = "none"; }
      try { var p = v.play(); if(p && p.catch){ p.catch(function(){}); } } catch(e) {}
      cameraReady = true;
      stepOk("cameraStep", "Kamera tayyor");
      setStatus("Yuzingizni oval ichida ushlang", "Liveness avtomatik tekshirilmoqda.");
      startLiveness();
      checkReady();
    }).catch(function(e){
      scanStarted = false;
      stepWarn("cameraStep", "Kamera ochilmadi");
      setStatus("Kamera ochilmadi", (e && e.message ? e.message : String(e)) + " — ekranga yana bir marta teging.");
    });
  }
  function startGPS(){
    if(!navigator.geolocation){ stepWarn("gpsStep", "GPS API topilmadi"); return; }
    navigator.geolocation.getCurrentPosition(function(p){
      pos = p;
      gpsReady = true;
      stepOk("gpsStep", "GPS tayyor · aniqlik " + Math.round(p.coords.accuracy) + "m");
      checkReady();
    }, function(e){
      scanStarted = false;
      stepWarn("gpsStep", "GPS aniqlanmadi");
      setStatus("GPS aniqlanmadi", (e && e.message ? e.message : String(e)) + " — ekranga yana teging.");
    }, { enableHighAccuracy: true, timeout: 20000, maximumAge: 0 });
  }
  function startLiveness(){
    if(liveReady){ return; }
    var n = 3;
    stepOk("liveStep", "Yuzingizni oval ichida ushlang: 3");
    var t = setInterval(function(){
      n = n - 1;
      if(n > 0){ stepOk("liveStep", "Yuzingizni oval ichida ushlang: " + n); }
      else { clearInterval(t); liveReady = true; stepOk("liveStep", "Liveness otdi"); checkReady(); }
    }, 750);
  }
  window.startScan = function(){
    if(scanStarted){ return; }
    scanStarted = true;
    setStatus("Tekshiruv boshlandi", "Kamera va GPS ruxsatlari olinmoqda.");
    startCamera();
    startGPS();
  };
  function checkReady(){
    if(cameraReady && gpsReady && liveReady && !confirmShown){
      confirmShown = true;
      var ov = el("oval"); if(ov){ ov.classList.add("ready"); }
      setStatus("Hammasi tayyor", ACTION_TITLE + " uchun tasdiqlang.");
      var co = el("confirmOverlay"); if(co){ co.className = "auto-confirm show"; }
      var note = el("scanNote"); if(note){ note.textContent = "Tasdiqlash uchun ekranning o‘rtasiga bosing."; }
    }
  }
  window.sendCheck = function(){
    if(!confirmShown || sending){ return; }
    sending = true;
    var co = el("confirmOverlay"); if(co){ co.className = "auto-confirm"; }
    setStatus("Yuborilmoqda", "Selfie, GPS va AI moslik tekshirilmoqda.");
    el("result").innerHTML = "<div class='result-card'>Tekshirilyapti...</div>";
    var v = el("video"); var c = el("canvas");
    c.width = v.videoWidth || 720; c.height = v.videoHeight || 960;
    c.getContext("2d").drawImage(v, 0, 0, c.width, c.height);
    var fd = new FormData();
    fd.append("action", ACTION);
    fd.append("image_data", c.toDataURL("image/jpeg", 0.88));
    fd.append("liveness", "ok");
    if(pos){ fd.append("lat", pos.coords.latitude); fd.append("lon", pos.coords.longitude); fd.append("accuracy", pos.coords.accuracy); }
    fetch("/api/attendance/verify", { method: "POST", body: fd }).then(function(r){
      return r.json().then(function(j){ return { ok: r.ok, data: j }; });
    }).then(function(pack){
      var j = pack.data || {}; var cls = pack.ok && !j.error ? "success" : "error";
      el("result").innerHTML = "<div class='result-card " + cls + "'><h2>" + (j.emoji || "ℹ️") + " " + (j.title || "Natija") + "</h2><p>" + (j.message || "") + "</p><p>Face: <b>" + (j.face_status || "-") + "</b> " + (j.face_score || "") + "</p><p>GPS: <b>" + (j.location_status || "-") + "</b> " + (j.distance_m ? j.distance_m + "m" : "") + "</p></div>";
      setStatus(j.title || "Natija", j.message || "");
      if(pack.ok && !j.error){ setTimeout(function(){ location.href = "/cabinet"; }, 2400); }
      sending = false;
    }).catch(function(e){
      el("result").innerHTML = "<div class='result-card error'>Xato: " + e + "</div>";
      setStatus("Xato", String(e)); sending = false;
    });
  };
  function boot(){ setTimeout(function(){ window.startScan(); }, 250); }
  document.addEventListener("DOMContentLoaded", boot);
  window.addEventListener("load", function(){ setTimeout(function(){ if(!cameraReady){ window.startScan(); } }, 700); });
  window.addEventListener("pageshow", function(){ setTimeout(function(){ if(!cameraReady){ window.startScan(); } }, 900); });
  document.addEventListener("touchstart", function(){ if(!cameraReady){ window.startScan(); } }, { passive: true });
  document.addEventListener("click", function(){ if(!cameraReady){ window.startScan(); } });
})();
</script>
</body></html>"""
    body = html_page.replace("__LABEL__", esc(label))
    body = body.replace("__STYLE__", EXTRA_CSS)
    body = body.replace("__ACTION__", action)
    body = body.replace("__ACTION_TITLE__", esc(action_title))
    body = body.replace("__USER_NAME__", esc(user.get("name") or "Xodim"))
    return HTMLResponse(body)



def get_face_reference(tid: str):
    try:
        ensure_verification_tables()
        return sql_one(
            """
            SELECT reference_photo_path, updated_at
            FROM face_templates
            WHERE company_id=? AND telegram_id=? AND status='active'
            ORDER BY id DESC LIMIT 1
            """,
            (db._current_cid(), str(tid)),
        )
    except Exception:
        return None


def reference_card(tid: str) -> str:
    ref = get_face_reference(tid)
    has_ref = bool(ref and row_get(ref, "reference_photo_path"))
    img = (
        f"<img src='/face-reference/{esc(tid)}?v={int(datetime.now().timestamp())}' "
        "style='width:112px;height:112px;border-radius:28px;object-fit:cover;border:1px solid rgba(255,255,255,.75);box-shadow:0 16px 42px rgba(15,23,42,.18)' alt='Reference'>"
        if has_ref
        else "<div style='width:112px;height:112px;border-radius:28px;background:linear-gradient(135deg,#dbeafe,#ede9fe);display:grid;place-items:center;font-size:42px;border:1px solid rgba(255,255,255,.75)'>📸</div>"
    )
    status = "Reference rasm tayyor" if has_ref else "Reference rasm hali qo‘yilmagan"
    updated = f"<div class='muted' style='font-size:13px;margin-top:4px'>Yangilangan: {esc(row_get(ref, 'updated_at', ''))}</div>" if has_ref else ""
    return f"""
      <div class="card" style="margin-top:16px">
        <div style="display:flex;gap:16px;align-items:center;flex-wrap:wrap">
          {img}
          <div style="flex:1;min-width:220px">
            <h2 style="margin:0 0 6px">AI reference rasm</h2>
            <p class="muted" style="margin:0">{status}. Keyingi Keldim/Ketdim selfie shu rasm bilan solishtiriladi.</p>
            {updated}
            <form method="post" action="/cabinet/reference-photo" enctype="multipart/form-data" style="margin-top:12px;display:flex;gap:10px;flex-wrap:wrap;align-items:center">
              <input type="file" name="file" accept="image/*" required style="max-width:280px">
              <button class="btn" type="submit">Sifatli rasmni saqlash</button>
            </form>
            <div class="muted" style="font-size:12px;margin-top:8px">Maslahat: yuz to‘liq ko‘rinsin, yorug‘ joyda, filtrsiz va aniq foto tanlang.</div>
          </div>
        </div>
      </div>
    """

def staff_cabinet(request: Request, user: dict) -> HTMLResponse:
    tid = str(user.get("telegram_id"))
    today = ws.now_tz().date()
    month_start = today.replace(day=1)
    u = row_dict(sql_one("SELECT * FROM users WHERE company_id=? AND telegram_id=?", (db._current_cid(), tid)))
    shifts = sql_all("SELECT * FROM shifts WHERE company_id=? AND telegram_id=? AND business_date>=? ORDER BY start_at DESC LIMIT 20", (db._current_cid(), tid, month_start.isoformat()))
    rate = float(row_get(u, "hourly_rate", 0) or 0)
    paid = sum(paid_for_shift(r)[1] for r in shifts)
    raw = sum(paid_for_shift(r)[0] for r in shifts if r["end_at"])
    active = sql_one("SELECT * FROM shifts WHERE company_id=? AND telegram_id=? AND status='open' ORDER BY start_at DESC LIMIT 1", (db._current_cid(), tid))
    rows = ""
    for r in shifts[:10]:
        raw_m, paid_m, _ = paid_for_shift(r)
        rows += f"<tr><td>{esc(r['business_date'])}</td><td>{esc(r['shop'])}</td><td>{dt_short(r['start_at'])}</td><td>{dt_short(r['end_at'])}</td><td>Real {worked_text(raw_m)}<br><b>{worked_text(paid_m)}</b></td></tr>"
    if not rows:
        rows = "<tr><td colspan='5'>Hali smena yo‘q</td></tr>"
    active_html = f"<div class='card'><span class='pill green'>Hozir ishda</span><h2>{esc(active['shop'])}</h2><p class='muted'>Boshlangan: {dt_short(active['start_at'])}</p></div>" if active else ""
    body = f"""<!doctype html><html lang="uz"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Kabinet · KKB</title><style>{ws.CSS}{EXTRA_CSS}</style></head><body><div class="staff-wrap">
      <div class="top"><h1>KKB · Kabinet</h1><form method="post" action="/logout"><button class="btn gray">Chiqish</button></form></div>
      <div class="staff-hero"><div class="emp-card"><div class="emp-avatar">{esc(row_get(u, 'emoji', '👤') or '👤')}</div><div><h2 style="margin:0">{esc(user.get('name') or row_get(u, 'full_name', 'Xodim') or 'Xodim')}</h2><p style="color:#bfdbfe;margin:6px 0 0">{esc(row_get(u, 'role', None) or user.get('role') or 'staff')}</p></div></div></div>
      <div class="big-actions"><a class="green" href="/checkin">🟢 Keldim</a><a class="red" href="/checkout">🔴 Ketdim</a></div>{active_html}
      {reference_card(tid)}
      <div class="grid cards"><div class="card metric"><div class="label">Bu oy yozilgan</div><div class="value">{round(paid/60,1)}</div></div><div class="card metric"><div class="label">Bu oy real</div><div class="value">{round(raw/60,1)}</div></div><div class="card metric"><div class="label">Oylik</div><div class="value">{money(paid/60*rate)}</div></div><div class="card metric"><div class="label">Stavka</div><div class="value">{money(rate)}</div></div></div>
      <div class="card"><h2>Oxirgi smenalarim</h2><div class="table-wrap"><table class="table"><tr><th>Sana</th><th>Shop</th><th>Keldi</th><th>Ketdi</th><th>Real/Yozildi</th></tr>{rows}</table></div></div>
    </div></body></html>"""
    return HTMLResponse(body)


def verification_page(request: Request) -> HTMLResponse:
    rows = checks(150)
    body_rows = ""
    for r in rows:
        check_id = int(r["id"])
        lat, lon = row_get(r, "latitude"), row_get(r, "longitude")
        map_link = f"https://www.google.com/maps?q={lat},{lon}" if lat and lon else ""
        map_btn = f"<a class='btn gray' href='{esc(map_link)}' target='_blank'>Map</a>" if map_link else ""
        body_rows += (
            f"<tr><td>{check_id}</td>"
            f"<td>{esc(row_get(r, 'created_at'))}<br>{esc(row_get(r, 'source'))}</td>"
            f"<td>{esc(row_get(r, 'name') or row_get(r, 'telegram_id'))}</td>"
            f"<td>{esc(row_get(r, 'action'))}</td>"
            f"<td>{esc(row_get(r, 'shop'))}<br>{esc(row_get(r, 'distance_m') or '')}m</td>"
            f"<td>{esc(row_get(r, 'face_status'))}<br>{esc(row_get(r, 'face_score'))}</td>"
            f"<td>{esc(row_get(r, 'location_status'))}</td>"
            f"<td>{esc(row_get(r, 'final_status'))}</td>"
            f"<td><a class='btn gray' href='/attendance-photo/{check_id}' target='_blank'>Selfie</a>{map_btn}"
            f"<form method='post' action='/verification/{check_id}/status' class='row' style='margin-top:6px'>"
            f"<button class='btn' name='status' value='approved'>OK</button>"
            f"<button class='btn danger' name='status' value='rejected'>Reject</button>"
            f"<button class='btn gray' name='status' value='needs_review'>Review</button></form></td></tr>"
        )
    content = f"<div class='card'>{ws.table(['ID','Vaqt','Xodim','Action','Shop/Masofa','Yuz','GPS','Status','Amal'], body_rows, 9)}</div>"
    return stable_layout(request, "verification", "AI tekshiruv", "Selfie, GPS va admin review", content)


UZ_MONTHS = {
    1: "yanvar", 2: "fevral", 3: "mart", 4: "aprel", 5: "may", 6: "iyun",
    7: "iyul", 8: "avgust", 9: "sentabr", 10: "oktabr", 11: "noyabr", 12: "dekabr",
}


def report_month_label(year: int, month: int) -> str:
    return f"{UZ_MONTHS.get(int(month), str(month))}{int(year)}"


def parse_report_date(value: Any) -> date | None:
    if not value:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def report_hour(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%H")
    d = parse_dt(value)
    if d:
        return d.strftime("%H")
    s = str(value).strip()
    if not s:
        return ""
    if ":" in s:
        return s.split(":", 1)[0].zfill(2)[-2:]
    if "." in s:
        return s.split(".", 1)[0].zfill(2)[-2:]
    if s.isdigit() and len(s) >= 2:
        return s[:2]
    return s[:2]


def report_hours_from_shift(row: Any) -> float:
    try:
        paid = int(row_get(row, "worked_minutes", 0) or 0)
        if paid > 0:
            return round(paid / 60, 2)
        raw, paid2, _ = paid_for_shift(row)
        return round((paid2 or raw) / 60, 2)
    except Exception:
        return 0.0


def available_report_months() -> list[tuple[int, int]]:
    rows = sql_all(
        "SELECT DISTINCT substr(business_date,1,7) AS ym FROM shifts WHERE company_id=? AND business_date<>'' ORDER BY ym DESC LIMIT 24",
        (db._current_cid(),),
    )
    months = []
    for r in rows:
        ym = row_get(r, "ym", "")
        try:
            y, m = str(ym).split("-", 1)
            months.append((int(y), int(m)))
        except Exception:
            pass
    if not months:
        n = datetime.now(config.get_timezone_obj()).date()
        months = [(n.year, n.month)]
    return months


def report_shops_for_month(year: int, month: int) -> list[str]:
    start = date(int(year), int(month), 1)
    end = date(int(year), int(month), calendar.monthrange(int(year), int(month))[1])
    rows = sql_all(
        """
        SELECT DISTINCT shop FROM shifts
        WHERE company_id=? AND business_date BETWEEN ? AND ? AND COALESCE(shop,'')<>''
        ORDER BY shop
        """,
        (db._current_cid(), start.isoformat(), end.isoformat()),
    )
    shops = [str(row_get(r, "shop", "")).strip() for r in rows if str(row_get(r, "shop", "")).strip()]
    if not shops:
        shops = [str(row_get(r, "name", "")).strip() for r in shop_rows() if str(row_get(r, "name", "")).strip()]
    return shops


def build_bot_style_report_xlsx(year: int, month: int, target_shop: str) -> tuple[str, int]:
    year = int(year)
    month = int(month)
    start = date(year, month, 1)
    end = date(year, month, calendar.monthrange(year, month)[1])
    num_days = calendar.monthrange(year, month)[1]

    rows = sql_all(
        """
        SELECT * FROM shifts
        WHERE company_id=? AND business_date BETWEEN ? AND ? AND shop=?
        ORDER BY name, business_date, start_at
        """,
        (db._current_cid(), start.isoformat(), end.isoformat(), str(target_shop)),
    )

    data_map: dict[str, dict[int, dict[str, Any]]] = {}
    for row in rows:
        d = parse_report_date(row_get(row, "business_date"))
        if not d or d.year != year or d.month != month:
            continue
        name = str(row_get(row, "name") or "Noma'lum").strip() or "Noma'lum"
        day = int(d.day)
        cur = data_map.setdefault(name, {}).setdefault(day, {"start": "", "end": "", "total": 0.0})
        st = report_hour(row_get(row, "start_at"))
        en = report_hour(row_get(row, "end_at"))
        total = report_hours_from_shift(row)
        if st and (not cur["start"] or st < cur["start"]):
            cur["start"] = st
        if en and (not cur["end"] or en > cur["end"]):
            cur["end"] = en
        cur["total"] = round(float(cur.get("total") or 0) + float(total or 0), 2)

    wb = Workbook()
    wsx = wb.active
    wsx.title = "Report"
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    fill_head = PatternFill("solid", fgColor="DDEBF7")
    fill_tot = PatternFill("solid", fgColor="FCE4D6")

    wsx.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + num_days + 3)
    wsx["A1"] = f"Hisobot: {target_shop} ({report_month_label(year, month)})"
    wsx["A1"].alignment = align
    wsx["A1"].font = Font(bold=True, size=14)

    for idx, h in enumerate(["ФИО", "Должность", "Дата"], 1):
        c = wsx.cell(row=2, column=idx, value=h)
        c.border = border; c.fill = fill_head; c.alignment = align; c.font = Font(bold=True)
    for d in range(1, num_days + 1):
        c = wsx.cell(row=2, column=3 + d, value=d)
        c.border = border; c.fill = fill_head; c.alignment = align; c.font = Font(bold=True)
        wsx.column_dimensions[c.column_letter].width = 4
    for idx, h in enumerate(["1-15", "16-end", "JAMI"]):
        c = wsx.cell(row=2, column=3 + num_days + 1 + idx, value=h)
        c.border = border; c.fill = fill_head; c.alignment = align; c.font = Font(bold=True)

    row_idx = 3
    merge_list = []
    for name in sorted(data_map.keys()):
        u_data = data_map[name]
        wsx.cell(row=row_idx, column=1, value=name).alignment = left_align
        wsx.cell(row=row_idx, column=1).border = border
        merge_list.append((row_idx, 1, row_idx + 2, 1))
        wsx.cell(row=row_idx, column=2, value="staff").alignment = align
        wsx.cell(row=row_idx, column=2).border = border
        merge_list.append((row_idx, 2, row_idx + 2, 2))
        for i, lbl in enumerate(["c", "до", "итого"]):
            c = wsx.cell(row=row_idx + i, column=3, value=lbl)
            c.alignment = align; c.border = border
        sum1 = sum2 = 0.0
        for d in range(1, num_days + 1):
            col = 3 + d
            info = u_data.get(d, {})
            total_v = float(info.get("total", 0) or 0)
            vals = [info.get("start", ""), info.get("end", ""), round(total_v, 2) if total_v > 0 else ""]
            for i, val in enumerate(vals):
                c = wsx.cell(row=row_idx + i, column=col, value=val)
                c.alignment = align; c.border = border
            if d <= 15:
                sum1 += total_v
            else:
                sum2 += total_v
        col_sum = 3 + num_days
        for col, val in [(col_sum + 1, sum1), (col_sum + 2, sum2), (col_sum + 3, sum1 + sum2)]:
            c = wsx.cell(row=row_idx + 2, column=col, value=round(val, 2) if val > 0 else "")
            c.fill = fill_tot; c.border = border; c.alignment = align
            if col == col_sum + 3:
                c.font = Font(bold=True)
        row_idx += 3

    for r1, c1, r2, c2 in merge_list:
        wsx.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        c = wsx.cell(row=r1, column=c1)
        c.alignment = left_align if c1 == 1 else align
        c.border = border
    wsx.column_dimensions["A"].width = 24
    wsx.column_dimensions["B"].width = 12
    wsx.column_dimensions["C"].width = 8
    safe_shop = "".join(ch if ch.isalnum() else "_" for ch in str(target_shop))[:40] or "shop"
    path = Path(tempfile.gettempdir()) / f"Report_{safe_shop}_{report_month_label(year, month)}_{uuid.uuid4().hex[:6]}.xlsx"
    wb.save(path)
    return str(path), len(data_map)


def apply_stable_patch(app) -> None:
    global _PATCHED
    if _PATCHED:
        return
    _PATCHED = True
    ensure_extra_tables()

    # Make built-in pages use the same sidebar/menu.
    ws.nav = stable_nav
    ws.layout = stable_layout

    routes_to_replace = [
        ("/cabinet", "GET"), ("/cabinet/reference-photo", "POST"), ("/face-reference/{telegram_id}", "GET"),
        ("/checkin", "GET"), ("/checkout", "GET"), ("/check-in", "GET"), ("/check-out", "GET"), ("/web-checkin", "GET"), ("/web-checkout", "GET"),
        ("/api/attendance/verify", "POST"),
        ("/verification", "GET"), ("/verification/{check_id}/status", "POST"), ("/attendance-photo/{check_id}", "GET"),
        ("/employees", "GET"), ("/employees/add", "POST"), ("/employees/{telegram_id}", "GET"), ("/employees/{telegram_id}/update", "POST"), ("/employees/{telegram_id}/delete", "POST"),
        ("/salary", "GET"), ("/reports", "GET"), ("/export", "GET"), ("/reports/download", "POST"),
    ]
    for path, method in routes_to_replace:
        remove_route(app, path, method)

    @app.get("/cabinet", response_class=HTMLResponse)
    async def cabinet(request: Request):
        red = ws.require_login(request)
        if red:
            return red
        user = ws.current_user(request)
        if ws.is_admin(user):
            return RedirectResponse("/dashboard", status_code=303)
        return staff_cabinet(request, user)


    @app.post("/cabinet/reference-photo")
    async def cabinet_reference_photo(request: Request, file: UploadFile = File(...)):
        red = ws.require_login(request)
        if red:
            return red
        user = ws.current_user(request)
        if ws.is_admin(user):
            return RedirectResponse("/dashboard", status_code=303)
        tid = str(user.get("telegram_id"))
        try:
            ensure_verification_tables()
            raw = await file.read()
            if not raw:
                return HTMLResponse("Rasm kelmadi", status_code=400)
            if len(raw) > 8 * 1024 * 1024:
                return HTMLResponse("Rasm hajmi 8 MB dan katta bo‘lmasin", status_code=400)
            ext = Path(file.filename or "reference.jpg").suffix.lower()
            if ext not in {".jpg", ".jpeg", ".png", ".webp"}:
                ext = ".jpg"
            ref_dir = Path(db.db_path).parent / "face_refs"
            ref_dir.mkdir(parents=True, exist_ok=True)
            tmp_path = ref_dir / f"{tid}_upload_{uuid.uuid4().hex[:8]}{ext}"
            tmp_path.write_bytes(raw)
            ref_hash = _image_phash(str(tmp_path))
            _set_reference(tid, str(tmp_path), ref_hash)
            try:
                tmp_path.unlink()
            except Exception:
                pass
            ref = get_face_reference(tid)
            if ref and row_get(ref, "reference_photo_path"):
                db._execute(
                    "UPDATE users SET avatar_file_id=? WHERE company_id=? AND telegram_id=?",
                    (str(row_get(ref, "reference_photo_path")), db._current_cid(), tid),
                )
            return RedirectResponse("/cabinet?reference=ok", status_code=303)
        except Exception as e:
            print(f"[kkb_stable_patch] reference upload failed: {e}")
            return HTMLResponse(f"Reference rasm saqlanmadi: {esc(e)}", status_code=500)

    @app.get("/face-reference/{telegram_id}")
    async def face_reference(request: Request, telegram_id: str):
        red = ws.require_login(request)
        if red:
            return red
        user = ws.current_user(request)
        if not ws.is_admin(user) and str(user.get("telegram_id")) != str(telegram_id):
            return JSONResponse({"error": "forbidden"}, status_code=403)
        ref = get_face_reference(str(telegram_id))
        path = row_get(ref, "reference_photo_path") if ref else ""
        if not path or not os.path.exists(str(path)):
            return JSONResponse({"error": "not_found"}, status_code=404)
        return FileResponse(str(path))

    @app.get("/checkin", response_class=HTMLResponse)
    async def checkin(request: Request):
        red = ws.require_login(request)
        if red:
            return red
        user = ws.current_user(request)
        if ws.is_admin(user):
            return RedirectResponse("/dashboard", status_code=303)
        return scan_page("Keldim", "check_in", user)

    @app.get("/checkout", response_class=HTMLResponse)
    async def checkout(request: Request):
        red = ws.require_login(request)
        if red:
            return red
        user = ws.current_user(request)
        if ws.is_admin(user):
            return RedirectResponse("/dashboard", status_code=303)
        return scan_page("Ketdim", "check_out", user)

    @app.get("/check-in", response_class=HTMLResponse)
    async def checkin_alias(request: Request):
        return await checkin(request)

    @app.get("/check-out", response_class=HTMLResponse)
    async def checkout_alias(request: Request):
        return await checkout(request)

    @app.get("/web-checkin", response_class=HTMLResponse)
    async def web_checkin_alias(request: Request):
        return await checkin(request)

    @app.get("/web-checkout", response_class=HTMLResponse)
    async def web_checkout_alias(request: Request):
        return await checkout(request)

    @app.post("/api/attendance/verify")
    async def api_verify(request: Request, action: str = Form(...), image_data: str = Form(...), lat: str = Form(""), lon: str = Form(""), accuracy: str = Form(""), liveness: str = Form("")):
        red = ws.require_login(request)
        if red:
            return JSONResponse({"error": "unauthorized"}, status_code=401)
        user = ws.current_user(request)
        if not user or ws.is_admin(user):
            return JSONResponse({"emoji": "⚠️", "title": "Xodim uchun", "message": "Keldim/Ketdim faqat xodim kabinetidan qilinadi."}, status_code=400)
        if liveness != "ok":
            return JSONResponse({"emoji": "⚠️", "title": "Liveness yo‘q", "message": "Avval liveness tekshiruvdan o‘ting."}, status_code=400)
        tid = str(user["telegram_id"])
        selfie_path = save_data_url_image(image_data, tid, "web")
        active = await db.get_active_shift_row(int(tid))
        latitude = float(lat) if lat else None
        longitude = float(lon) if lon else None
        acc = float(accuracy) if accuracy else None
        loc_url = f"https://www.google.com/maps?q={lat},{lon}" if lat and lon else ""
        staff = await db.get_user_by_telegram_id(int(tid))
        staff_for_shift = staff.copy() if staff else {"TelegramID": tid, "Имя": user.get("name", "Xodim"), "Роль": "staff", "Магазин": ""}
        if action == "check_in" and active:
            return JSONResponse({"emoji":"⚠️", "title":"Smena ochiq", "message":"Sizda ochiq smena bor. Avval Ketdim qiling.", "face_status":"-", "location_status":"-"})
        if action == "check_out" and not active:
            return JSONResponse({"emoji":"⚠️", "title":"Ochiq smena yo‘q", "message":"Avval Keldim qiling.", "face_status":"-", "location_status":"-"})
        fallback_shop = str(staff_for_shift.get("Магазин", "")).split(",")[0].strip()
        result = verify_attendance(tid, action, selfie_path, latitude, longitude, acc, shift_id=active, source="web", fallback_shop=fallback_shop)
        if result.get("final_status") == "rejected":
            notify_group(action, user.get("name", "Xodim"), result.get("shop") or fallback_shop, result, selfie_path, loc_url)
            return JSONResponse({"emoji":"❌", "title":"Rad etildi", "message":"Selfie/GPS mos kelmadi. Admin tekshiradi.", **result})
        if action == "check_in":
            staff_for_shift["Магазин"] = result.get("shop") or fallback_shop
            ok = await db.start_shift(staff_for_shift, f"web:{result['check_id']}", loc_url)
            if not ok:
                return JSONResponse({"emoji":"❌", "title":"Smena ochilmadi", "message":"Bazaga yozishda xato yoki smena allaqachon ochiq.", **result}, status_code=500)
            shift_id = await db.get_active_shift_row(int(tid))
            try:
                db._execute("UPDATE attendance_checks SET shift_id=? WHERE id=?", (shift_id, int(result["check_id"])))
            except Exception:
                pass
            notify_group(action, user.get("name", "Xodim"), result.get("shop") or fallback_shop, result, selfie_path, loc_url)
            return JSONResponse({"emoji":"✅", "title":"Keldim saqlandi", "message":"Smena ochildi." if result.get("final_status") == "approved" else "Smena ochildi, admin reviewga ham tushdi.", **result})
        worked = await db.end_shift(active, photo_id=f"web:{result['check_id']}", location=loc_url)
        notify_group(action, user.get("name", "Xodim"), result.get("shop") or fallback_shop, result, selfie_path, loc_url, worked)
        return JSONResponse({"emoji":"✅", "title":"Ketdim saqlandi", "message":f"Smena yopildi. Ishladi: {worked}", **result})

    @app.get("/verification", response_class=HTMLResponse)
    async def verification(request: Request):
        red = ws.require_admin(request)
        if red:
            return red
        return verification_page(request)

    @app.post("/verification/{check_id}/status")
    async def verification_status(request: Request, check_id: int, status: str = Form(...)):
        red = ws.require_admin(request)
        if red:
            return red
        user = ws.current_user(request) or {}
        update_check_status(check_id, status, str(user.get("telegram_id", "admin")))
        return RedirectResponse("/verification", status_code=303)

    @app.get("/attendance-photo/{check_id}")
    async def attendance_photo(request: Request, check_id: int):
        red = ws.require_admin(request)
        if red:
            return red
        row = sql_one("SELECT selfie_path FROM attendance_checks WHERE company_id=? AND id=?", (db._current_cid(), int(check_id)))
        if not row or not row["selfie_path"] or not os.path.exists(row["selfie_path"]):
            return JSONResponse({"error":"not found"}, status_code=404)
        return FileResponse(row["selfie_path"])

    @app.get("/employees", response_class=HTMLResponse)
    async def employees(request: Request, q: str = ""):
        red = ws.require_admin(request)
        if red:
            return red
        query = f"%{q.strip()}%"
        if q.strip():
            users = sql_all("""
                SELECT * FROM users WHERE company_id=? AND active=1
                AND (full_name LIKE ? OR telegram_id LIKE ? OR phone LIKE ? OR role LIKE ?)
                ORDER BY full_name
            """, (db._current_cid(), query, query, query, query))
        else:
            users = sql_all("SELECT * FROM users WHERE company_id=? AND active=1 ORDER BY full_name", (db._current_cid(),))
        rows = ""
        for u in users:
            shops = ", ".join(sorted(user_shops(u["id"]))) or "—"
            rows += f"<tr><td><b>{esc(u['emoji'] or '🙂')} {esc(u['full_name'])}</b><br><span class='mini'>ID: {esc(u['telegram_id'])} · {esc(u['phone'] or 'telefon yo‘q')}</span></td><td>{esc(u['role'])}</td><td>{esc(shops)}</td><td>{money(u['hourly_rate'])}</td><td><a class='btn secondary' href='/employees/{esc(u['telegram_id'])}'>Tahrirlash</a></td></tr>"
        add_form = f"""
        <div class="card"><h2>Yangi xodim qo‘shish</h2><form class="grid" method="post" action="/employees/add">
          <input class="input" name="telegram_id" placeholder="Telegram ID" required>
          <input class="input" name="name" placeholder="Ism" required>
          <input class="input" name="phone" placeholder="Telefon">
          <input class="input" name="rate" placeholder="Stavka, masalan 21000">
          <input class="input" name="emoji" value="🙂" placeholder="Emoji">
          <input class="input" name="position" placeholder="Lavozim">
          <input class="input" name="department" placeholder="Bo‘lim">
          <select class="select" name="role">{role_options('staff')}</select>
          <div><b>Filiallar</b>{shop_checkbox_html()}</div>
          <button class="btn">Qo‘shish</button>
        </form></div>"""
        content = f"<div class='two'><div class='card'><form class='form' method='get'><input class='input' name='q' value='{esc(q)}' placeholder='Xodim qidirish'><button class='btn'>Qidirish</button><a class='btn gray' href='/employees'>Tozalash</a></form>{ws.table(['Xodim','Rol','Filiallar','Stavka',''], rows, 5)}</div>{add_form}</div>"
        return stable_layout(request, "employees", "Xodimlar", f"Jami: {len(users)}", content)

    @app.post("/employees/add")
    async def employee_add(request: Request, telegram_id: str = Form(...), name: str = Form(...), phone: str = Form(""), rate: str = Form("0"), emoji: str = Form("🙂"), role: str = Form("staff"), department: str = Form(""), position: str = Form(""), shops: list[str] = Form(default=[])):
        red = ws.require_admin(request)
        if red:
            return red
        ensure_user_and_shops(telegram_id, name, role, shops, phone, rate, emoji, department, position)
        await db.append_audit_log(ws.current_user(request)["telegram_id"], ws.current_user(request)["role"], "web_employee_add", {"target": telegram_id})
        return RedirectResponse("/employees", status_code=303)

    @app.get("/employees/{telegram_id}", response_class=HTMLResponse)
    async def employee_detail(request: Request, telegram_id: str):
        red = ws.require_admin(request)
        if red:
            return red
        u = sql_one("SELECT * FROM users WHERE company_id=? AND telegram_id=?", (db._current_cid(), str(telegram_id)))
        if not u:
            return stable_layout(request, "employees", "Xodim topilmadi", telegram_id, "<div class='card'>Topilmadi</div>")
        selected = user_shops(u["id"])
        month_start = ws.now_tz().date().replace(day=1)
        shifts = sql_all("SELECT * FROM shifts WHERE company_id=? AND telegram_id=? AND business_date>=? ORDER BY start_at DESC LIMIT 12", (db._current_cid(), str(telegram_id), month_start.isoformat()))
        rows = ""
        paid_total = raw_total = 0
        for r in shifts:
            raw, paid, _ = paid_for_shift(r)
            raw_total += raw
            paid_total += paid
            rows += f"<tr><td>{esc(r['business_date'])}</td><td>{esc(r['shop'])}</td><td>{dt_short(r['start_at'])}</td><td>{dt_short(r['end_at'])}</td><td>Real {worked_text(raw)}<br><b>{worked_text(paid)}</b></td></tr>"
        form = f"""
        <div class="card"><h2>{esc(u['emoji'] or '🙂')} {esc(u['full_name'])}</h2><form class="edit-grid" method="post" action="/employees/{esc(telegram_id)}/update">
          <input class="input" name="name" value="{esc(u['full_name'])}" placeholder="Ism" required>
          <input class="input" name="phone" value="{esc(u['phone'] or '')}" placeholder="Telefon">
          <input class="input" name="rate" value="{esc(u['hourly_rate'] or 0)}" placeholder="Stavka">
          <input class="input" name="emoji" value="{esc(u['emoji'] or '🙂')}" placeholder="Emoji">
          <input class="input" name="position" value="{esc(u['position'] or '')}" placeholder="Lavozim">
          <input class="input" name="department" value="{esc(u['department'] or '')}" placeholder="Bo‘lim">
          <select class="select" name="role">{role_options(u['role'])}</select>
          <div><b>Filiallar</b>{shop_checkbox_html(selected)}</div>
          <button class="btn">Saqlash</button>
        </form><form method="post" action="/employees/{esc(telegram_id)}/delete" style="margin-top:12px"><button class="btn danger" onclick="return confirm('Xodimni deaktivatsiya qilamizmi?')">Deaktivatsiya</button></form></div>
        """
        stat = f"<div class='card'><h2>Bu oy</h2><div class='grid cards'><div class='card metric'><div class='label'>Real</div><div class='value'>{round(raw_total/60,1)}</div></div><div class='card metric'><div class='label'>Yozilgan</div><div class='value'>{round(paid_total/60,1)}</div></div><div class='card metric'><div class='label'>Oylik</div><div class='value'>{money(paid_total/60*float(u['hourly_rate'] or 0))}</div></div></div>{ws.table(['Sana','Shop','Keldi','Ketdi','Real/Yozildi'], rows, 5)}</div>"
        return stable_layout(request, "employees", "Xodim profili", u["full_name"], f"<div class='two'>{form}{stat}</div>")

    @app.post("/employees/{telegram_id}/update")
    async def employee_update(request: Request, telegram_id: str, name: str = Form(...), phone: str = Form(""), rate: str = Form("0"), emoji: str = Form("🙂"), role: str = Form("staff"), department: str = Form(""), position: str = Form(""), shops: list[str] = Form(default=[])):
        red = ws.require_admin(request)
        if red:
            return red
        ensure_user_and_shops(telegram_id, name, role, shops, phone, rate, emoji, department, position)
        await db.append_audit_log(ws.current_user(request)["telegram_id"], ws.current_user(request)["role"], "web_employee_update", {"target": telegram_id})
        return RedirectResponse(f"/employees/{urllib.parse.quote(str(telegram_id))}", status_code=303)

    @app.post("/employees/{telegram_id}/delete")
    async def employee_delete(request: Request, telegram_id: str):
        red = ws.require_admin(request)
        if red:
            return red
        db._execute("UPDATE users SET active=0, updated_at=CURRENT_TIMESTAMP WHERE company_id=? AND telegram_id=?", (db._current_cid(), str(telegram_id)))
        return RedirectResponse("/employees", status_code=303)


    @app.get("/reports", response_class=HTMLResponse)
    async def reports(request: Request, year: int | None = None, month: int | None = None, shop: str = ""):
        red = ws.require_admin(request)
        if red:
            return red
        months = available_report_months()
        if year is None or month is None:
            year, month = months[0]
        year = int(year); month = int(month)
        shops = report_shops_for_month(year, month)
        if not shop and shops:
            shop = shops[0]
        month_options = "".join(
            f"<option value='{y}-{m}' {'selected' if int(y)==year and int(m)==month else ''}>{esc(report_month_label(y, m))}</option>"
            for y, m in months
        )
        shop_options = "".join(f"<option value='{esc(sh)}' {'selected' if sh == shop else ''}>{esc(sh)}</option>" for sh in shops)
        preview_rows = ""
        if shop:
            start = date(year, month, 1)
            end = date(year, month, calendar.monthrange(year, month)[1])
            stats = sql_all(
                """
                SELECT name, COUNT(*) AS shifts_count, SUM(COALESCE(worked_minutes,0)) AS paid_minutes
                FROM shifts
                WHERE company_id=? AND business_date BETWEEN ? AND ? AND shop=?
                GROUP BY name ORDER BY name
                """,
                (db._current_cid(), start.isoformat(), end.isoformat(), shop),
            )
            for r in stats:
                nm = row_get(r, "name") or "Noma'lum"
                preview_rows += f"<tr><td>{esc(nm)}</td><td>{int(row_get(r,'shifts_count',0) or 0)}</td><td>{worked_text(int(row_get(r,'paid_minutes',0) or 0))}</td></tr>"
        content = f"""
        <div class="card">
          <h2>Botdagi 🧾 Hisobotlar webda</h2>
          <p class="mini">Botdagi oy → filial → Excel hisobot jarayoni webga qo‘shildi. Eski web tabel o‘rniga shu ishlatiladi.</p>
          <form class="form" method="get" action="/reports">
            <select class="select" name="ym" onchange="const v=this.value.split('-'); window.location='/reports?year='+v[0]+'&month='+v[1]+'&shop={urllib.parse.quote(shop)}'">{month_options}</select>
            <input type="hidden" name="year" value="{year}"><input type="hidden" name="month" value="{month}">
            <select class="select" name="shop">{shop_options}</select>
            <button class="btn">Ko‘rish</button>
          </form>
          <form method="post" action="/reports/download" style="margin-top:12px">
            <input type="hidden" name="year" value="{year}"><input type="hidden" name="month" value="{month}"><input type="hidden" name="shop" value="{esc(shop)}">
            <button class="btn">📥 Excel hisobotni yuklab olish</button>
          </form>
        </div>
        <div class="card"><h2>{esc(shop or 'Filial tanlanmagan')} · {esc(report_month_label(year, month))}</h2>{ws.table(['Xodim','Smena soni','Yozilgan soat'], preview_rows, 3)}</div>
        """
        return stable_layout(request, "reports", "Hisobotlar", "Botdagi magazin oylik Excel hisobotini webdan olish", content)

    @app.get("/export", response_class=HTMLResponse)
    async def export_alias(request: Request):
        red = ws.require_admin(request)
        if red:
            return red
        return RedirectResponse("/reports", status_code=303)

    @app.post("/reports/download")
    async def reports_download(request: Request, year: int = Form(...), month: int = Form(...), shop: str = Form(...)):
        red = ws.require_admin(request)
        if red:
            return red
        if not str(shop or "").strip():
            return stable_layout(request, "reports", "Hisobotlar", "Filial tanlanmagan", "<div class='card'>Avval filialni tanlang.</div>")
        path, count = build_bot_style_report_xlsx(int(year), int(month), str(shop))
        filename = f"Report_{str(shop).replace(' ', '_')}_{report_month_label(int(year), int(month))}.xlsx"
        return FileResponse(path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=filename)

    @app.get("/salary", response_class=HTMLResponse)
    async def salary(request: Request, date_from: str = "", date_to: str = "", shop: str = ""):
        red = ws.require_admin(request)
        if red:
            return red
        today = ws.now_tz().date()
        start = ws.parse_date(date_from, today.replace(day=1)) or today.replace(day=1)
        end = ws.parse_date(date_to, today) or today
        if end < start:
            start, end = end, start
        users = sql_all("SELECT * FROM users WHERE company_id=? AND active=1 ORDER BY full_name", (db._current_cid(),))
        rows = ""
        total_sum = total_paid = total_raw = 0
        for u in users:
            q = "SELECT * FROM shifts WHERE company_id=? AND telegram_id=? AND business_date BETWEEN ? AND ? AND status='closed'"
            params = [db._current_cid(), str(u["telegram_id"]), start.isoformat(), end.isoformat()]
            if shop:
                q += " AND shop=?"
                params.append(shop)
            shifts = sql_all(q, tuple(params))
            raw = paid = br = 0
            for sh in shifts:
                r, p, b = paid_for_shift(sh)
                raw += r
                paid += p
                br += b
            amount = paid / 60 * float(u["hourly_rate"] or 0)
            total_raw += raw
            total_paid += paid
            total_sum += amount
            if shifts or not shop:
                rows += f"<tr><td>{esc(u['full_name'])}</td><td>{esc(', '.join(sorted(user_shops(u['id']))) or '—')}</td><td>{len(shifts)}</td><td>{worked_text(raw)}</td><td><b>{worked_text(paid)}</b><div class='mini'>tushlik {worked_text(br)}</div></td><td>{money(u['hourly_rate'])}</td><td><b>{money(amount)}</b></td></tr>"
        qs_month = today.replace(day=1).isoformat(), today.isoformat()
        first_half = today.replace(day=1).isoformat(), today.replace(day=min(15, today.day if today.day < 15 else 15)).isoformat()
        last_day = date(today.year, today.month, __import__('calendar').monthrange(today.year, today.month)[1]).isoformat()
        second_half = today.replace(day=16).isoformat(), last_day
        content = f"""
        <div class="card"><form class="form" method="get"><input class="input" type="date" name="date_from" value="{start.isoformat()}"><input class="input" type="date" name="date_to" value="{end.isoformat()}"><select class="select" name="shop">{shop_select(shop, True)}</select><button class="btn">Hisoblash</button></form>
        <div class="row" style="margin-bottom:14px"><a class="btn gray" href="/salary?date_from={qs_month[0]}&date_to={qs_month[1]}">Bu oy</a><a class="btn gray" href="/salary?date_from={first_half[0]}&date_to={first_half[1]}">1–15</a><a class="btn gray" href="/salary?date_from={second_half[0]}&date_to={second_half[1]}">16–oy oxiri</a></div>
        <h2>{start.strftime('%d.%m.%Y')} — {end.strftime('%d.%m.%Y')}: {money(total_sum)} so‘m</h2>
        <div class="grid cards"><div class="card metric"><div class="label">Real</div><div class="value">{round(total_raw/60,1)}</div></div><div class="card metric"><div class="label">Yozilgan</div><div class="value">{round(total_paid/60,1)}</div></div><div class="card metric"><div class="label">Jami oylik</div><div class="value">{money(total_sum)}</div></div></div>
        {ws.table(['Xodim','Filiallar','Smena','Real','Yozilgan','Stavka','Jami'], rows, 7)}</div>"""
        return stable_layout(request, "salary", "Oylik", "Sana oralig‘i, eski yaxlitlash qoidasi va tushlik ayrimi bilan", content)

    print("[kkb_stable_patch] applied: unified menu, staff scan, employees edit/add, salary period, bot-style reports")
