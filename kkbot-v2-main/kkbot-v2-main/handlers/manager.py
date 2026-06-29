import os
import calendar
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from aiogram import Router, F
from aiogram.types import Message, FSInputFile, CallbackQuery
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.utils.keyboard import InlineKeyboardBuilder

from database.sqlite_db import db

router = Router()
router.message.filter(F.chat.type == "private")

class ManagerReportState(StatesGroup):
    waiting_for_month = State()
    waiting_for_shop = State()

# --- OY NOMLARI (UZ) ---
# UI format: "dekabr2025", "yanvar2026" (bo'shliqsiz)
UZ_MONTHS = {
    1: "yanvar",
    2: "fevral",
    3: "mart",
    4: "aprel",
    5: "may",
    6: "iyun",
    7: "iyul",
    8: "avgust",
    9: "sentabr",
    10: "oktabr",
    11: "noyabr",
    12: "dekabr",
}

def month_label(year: int, month: int) -> str:
    return f"{UZ_MONTHS.get(month, str(month))}{year}"

async def get_available_months_from_shifts() -> list[tuple[int, int]]:
    """SQLite/Sмены ichidan mavjud (yil, oy) larni chiqaradi."""
    ws = await db._get_worksheet("Смены")
    dates = await ws.col_values(1)
    months_set: set[tuple[int, int]] = set()
    for i in range(1, len(dates)):  # 0 - header
        d_obj = parse_sheet_date(dates[i])
        if not d_obj:
            continue
        months_set.add((d_obj.year, d_obj.month))

    months = sorted(months_set, key=lambda x: (x[0], x[1]), reverse=True)
    return months


async def get_report_shops(year: int, month: int) -> list[str]:
    """Hisobot uchun shoplarni chiqaradi.

    Avval tanlangan oy ichidagi smenalardan distinct shop oladi. Agar current
    company_id/sheet_id mismatch sabab bo'sh chiqsa, fallback qilib barcha
    shoplardan oladi. Bu migratsiyadan keyingi eng ko'p uchraydigan holatni yopadi.
    """
    shops: list[str] = []

    # 1) Tanlangan oy smenalaridan olish - report uchun eng to'g'ri manba.
    try:
        start_d = datetime(year, month, 1).date()
        last_day = calendar.monthrange(year, month)[1]
        end_d = datetime(year, month, last_day).date()
        shift_rows = await db.get_shift_rows(start_date=start_d, end_date=end_d)
        shops = sorted({str(r.get("shop", "")).strip() for r in shift_rows if str(r.get("shop", "")).strip()})
    except Exception:
        shops = []

    # 2) Agar bo'sh bo'lsa, normal shops jadvalidan olish.
    if not shops:
        try:
            shops = [str(x).strip() for x in await db.get_shops() if str(x).strip()]
        except Exception:
            shops = []

    # 3) Eng oxirgi fallback: company_id farq qilib qolsa ham, DB ichidagi barcha shoplar.
    if not shops:
        try:
            rows = db._execute("SELECT DISTINCT name FROM shops WHERE active=1 AND name<>'' ORDER BY name", (), "all")
            shops = [r["name"] for r in rows]
        except Exception:
            shops = []

    # 4) Yana fallback: shifts jadvalidagi barcha shoplar.
    if not shops:
        try:
            rows = db._execute("SELECT DISTINCT shop FROM shifts WHERE shop<>'' ORDER BY shop", (), "all")
            shops = [r["shop"] for r in rows]
        except Exception:
            shops = []

    return shops

# --- TOZALASH FUNKSIYALARI ---
def clean_hour(val):
    """Vaqtni tozalab faqat soatni oladi (09:00 -> 09)"""
    if not val: return ""
    s = str(val).strip()
    if len(s) < 2: return ""
    
    # 09:30 -> 09
    if ":" in s: return s.split(":")[0]
    # 09.30 -> 09
    if "." in s: return s.split(".")[0]
    
    # Agar 4 ta harf bo'lsa (0900)
    if s.isdigit() and len(s) == 4: return s[:2]
    
    return s[:2]

def parse_sheet_date(date_val):
    if not date_val: return None
    s = str(date_val).strip()
    for fmt in ["%d-%m-%Y", "%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"]:
        try:
            return datetime.strptime(s, fmt)
        except: continue
    return None

def parse_float(val):
    if not val: return 0
    s = str(val).strip().replace(",", ".")
    try:
        # 8 ч 30 м -> 8.5
        if "ч" in s:
            parts = s.split("ч")
            h = int(parts[0])
            m = 0
            if "м" in parts[1]:
                m = int(parts[1].replace("м","").strip())
            return h + (m/60)
        return float(s)
    except: return 0

@router.message(F.text == "🧾 Hisobotlar")
async def report_start(message: Message, state: FSMContext, role: str):
    if role not in ["admin", "manager"]: return
    # 1) Avval oy tanlanadi
    try:
        months = await get_available_months_from_shifts()
    except Exception:
        months = []

    if not months:
        today = datetime.now()
        months = [(today.year, today.month)]

    builder = InlineKeyboardBuilder()
    for (y, m) in months[:24]:  # juda uzun bo'lib ketmasin
        builder.button(text=month_label(y, m), callback_data=f"rep_month_{y}_{m}")
    builder.adjust(3)

    await state.set_state(ManagerReportState.waiting_for_month)
    await message.answer("Qaysi oy uchun hisobot kerak?", reply_markup=builder.as_markup())


@router.callback_query(ManagerReportState.waiting_for_month, F.data.startswith("rep_month_"))
async def month_selected(callback: CallbackQuery, state: FSMContext, role: str):
    await callback.answer()
    if role not in ["admin", "manager"]:
        return

    try:
        _, _, y, m = callback.data.split("_", 3)
        year = int(y)
        month = int(m)
    except Exception:
        await callback.message.edit_text("Oy tanlashda xatolik. Qaytadan urinib ko'ring.")
        await state.clear()
        return

    await state.update_data(report_year=year, report_month=month)

    # 2) Keyin magazin tanlanadi. Shoplarni shu oy smenalaridan olamiz.
    shops = await get_report_shops(year, month)

    if not shops:
        await callback.message.edit_text(
            f"Tanlangan oy: <b>{month_label(year, month)}</b>\n\n"
            "Bu oy uchun magazinlar topilmadi. Migratsiya/audit scriptini tekshiring."
        )
        await state.clear()
        return

    shop_map = {str(i): shop for i, shop in enumerate(shops, start=1)}
    await state.update_data(report_shop_map=shop_map)

    builder = InlineKeyboardBuilder()
    for token, shop in shop_map.items():
        builder.button(text=shop, callback_data=f"rep_shop_{token}")
    builder.adjust(2)

    await state.set_state(ManagerReportState.waiting_for_shop)
    await callback.message.edit_text(
        f"Tanlangan oy: <b>{month_label(year, month)}</b>\n\nMagazinni tanlang:",
        reply_markup=builder.as_markup(),
    )

@router.callback_query(ManagerReportState.waiting_for_shop, F.data.startswith("rep_shop_"))
async def generate_report_v3(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    token = callback.data.split("rep_shop_")[1]
    st_data = await state.get_data()
    shop_map = st_data.get("report_shop_map", {})
    target_shop = shop_map.get(token, token)
    await callback.message.edit_text(f"⏳ {target_shop} hisoboti tayyorlanmoqda... (V3)")
    print(f"--- HISOBOT BOSHLANDI: {target_shop} ---")

    try:
        # 1. BAZADAN QATORLARNI OLISH (Values orqali)
        ws = await db._get_worksheet("Смены")
        rows = await ws.get_all_values() # List of Lists
        
        # Ustunlar indeksini aniqlash (A=0, B=1...)
        # A: Sana, B: ID, C: Ism, D: Magazin, E: Keldi, F: Ketdi, G: Otrabotano
        # Agar sarlavha 1-qatorda bo'lsa, ma'lumot 2-qatordan boshlanadi
        
        st_data = await state.get_data()
        year = int(st_data.get("report_year", datetime.now().year))
        month = int(st_data.get("report_month", datetime.now().month))
        num_days = calendar.monthrange(year, month)[1]
        
        data_map = {}

        # 1-qator bu Sarlavha, shuning uchun 1 dan boshlaymiz (index 0 emas)
        for i in range(1, len(rows)):
            row = rows[i]
            if len(row) < 7: continue # Chala qator

            # Indekslar (Google Sheetsdagi joylashuv bo'yicha)
            # 0=Sana, 1=ID, 2=Ism, 3=Magazin, 4=Keldi, 5=Ketdi, 6=Otrabotano
            
            r_date = row[0]
            r_name = row[2]
            r_shop = row[3]
            r_start = row[4]
            r_end = row[5]
            r_total = row[6]

            if str(r_shop).strip() != target_shop: continue
            
            d_obj = parse_sheet_date(r_date)
            if not d_obj: continue
            if d_obj.year != year or d_obj.month != month: continue

            print(f"Topildi: {r_name} | Kun: {d_obj.day} | {r_start}-{r_end}")

            name = str(r_name).strip()
            if not name: name = "Noma'lum"
            
            if name not in data_map:
                data_map[name] = {}
            
            day = d_obj.day
            data_map[name][day] = {
                'start': clean_hour(r_start),
                'end': clean_hour(r_end),
                'total': parse_float(r_total)
            }

        # 2. EXCEL YARATISH
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Format
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        align = Alignment(horizontal="center", vertical="center")
        fill_head = PatternFill("solid", fgColor="DDEBF7")
        fill_tot = PatternFill("solid", fgColor="FCE4D6")
        
        # Sarlavha
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3+num_days+3)
        ws['A1'] = f"Hisobot: {target_shop} ({month}.{year})"
        ws['A1'].alignment = align
        ws['A1'].font = Font(bold=True, size=14)

        # Shapka
        headers = ["ФИО", "Должность", "Дата"]
        for idx, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=idx, value=h)
            c.border = border
            c.fill = fill_head
            c.font = Font(bold=True)

        for d in range(1, num_days+1):
            c = ws.cell(row=2, column=3+d, value=d)
            c.border = border
            c.alignment = align
            c.fill = fill_head
            ws.column_dimensions[c.column_letter].width = 4

        # Jami ustunlari
        for idx, h in enumerate(["1-15", "16-end", "JAMI"]):
            c = ws.cell(row=2, column=3+num_days+1+idx, value=h)
            c.border = border
            c.fill = fill_head
            c.font = Font(bold=True)

        # Ma'lumot to'ldirish
        row_idx = 3
        sorted_names = sorted(data_map.keys())
        
        # Merge keyinroq qilinadi
        merge_list = []

        for name in sorted_names:
            u_data = data_map[name]
            
            # Ism (3 qator)
            ws.cell(row=row_idx, column=1, value=name).alignment = align
            ws.cell(row=row_idx, column=1).border = border
            merge_list.append((row_idx, 1, row_idx+2, 1))

            # Rol (3 qator)
            ws.cell(row=row_idx, column=2, value="staff").alignment = align
            ws.cell(row=row_idx, column=2).border = border
            merge_list.append((row_idx, 2, row_idx+2, 2))

            # Labels
            labels = ["c", "до", "итого"]
            for i, lbl in enumerate(labels):
                c = ws.cell(row=row_idx+i, column=3, value=lbl)
                c.alignment = align
                c.border = border

            # Kunlar
            sum1 = 0
            sum2 = 0
            
            for d in range(1, num_days+1):
                col = 3 + d
                info = u_data.get(d, {})
                
                # Start
                c1 = ws.cell(row=row_idx, column=col, value=info.get('start', ''))
                c1.alignment = align
                c1.border = border
                
                # End
                c2 = ws.cell(row=row_idx+1, column=col, value=info.get('end', ''))
                c2.alignment = align
                c2.border = border
                
                # Total
                val = info.get('total', 0)
                c3 = ws.cell(row=row_idx+2, column=col, value=val if val>0 else "")
                c3.alignment = align
                c3.border = border
                
                if d <= 15: sum1 += val
                else: sum2 += val

            # Jamilar
            col_sum = 3 + num_days
            
            c_s1 = ws.cell(row=row_idx+2, column=col_sum+1, value=sum1 if sum1>0 else "")
            c_s1.fill = fill_tot
            c_s1.border = border
            
            c_s2 = ws.cell(row=row_idx+2, column=col_sum+2, value=sum2 if sum2>0 else "")
            c_s2.fill = fill_tot
            c_s2.border = border
            
            c_s3 = ws.cell(row=row_idx+2, column=col_sum+3, value=sum1+sum2 if (sum1+sum2)>0 else "")
            c_s3.fill = fill_tot
            c_s3.border = border
            c_s3.font = Font(bold=True)

            row_idx += 3

        # Merge qilish (Oxirida)
        for (r1, c1, r2, c2) in merge_list:
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
            # Stilni to'g'irlash
            c = ws.cell(row=r1, column=c1)
            c.alignment = align
            c.border = border

        fname = f"Report_{target_shop}_{month_label(year, month)}.xlsx"
        wb.save(fname)
        await callback.message.answer_document(FSInputFile(fname))
        os.remove(fname)

    except Exception as e:
        await callback.message.answer(f"Xato: {e}")
        print(e)
    
    await state.clear()
